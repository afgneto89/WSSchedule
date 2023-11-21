import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
import datetime
from openpyxl import load_workbook

import smtplib
import email.message


def send_new_rows_email(new_rows, prod_name):
    email_content = f"Olá, um novo item foi inserido para monitoramento dos preços.\n\n<br>Produto: {prod_name}\n\n"

    for row in new_rows:
        email_content += f"<br>Min Quantity: {row['Min Quantity']}\nUnit: {row['Unit']}\nPrice: {row['Price']}\n\n"

    msg = email.message.Message()
    msg['Subject'] = 'Novo(s) item(ns) adicionado(s) ao monitoramento'
    msg['From'] = 'afneto@ajover.com'
    msg['To'] = 'afneto@ajover.com'
    password = 'Processos@23'
    msg.add_header('Content-Type', 'text/html', charset='utf-8')
    msg.set_payload(email_content, charset='utf-8')

    s = smtplib.SMTP('172.35.5.22:587')
    s.starttls()
    s.login(msg['From'], password)
    s.sendmail(msg['From'], [msg['To']], msg.as_string())


def send_changed_prices_email(changed_prices):
    email_content = "Preços foram alterados:\n"
    for row in changed_prices:
        email_content += f"Min Quantity: {row['Min Quantity']}, Price: {row['Price']}\n"

    msg = email.message.Message()
    msg['Subject'] = 'Preços foram alterados na planilha'
    msg['From'] = 'afneto@ajover.com'
    msg['To'] = 'afneto@ajover.com'
    password = 'Processos@23'
    msg.add_header('Content-Type', 'text/html')
    msg.set_payload(email_content)

    s = smtplib.SMTP('172.35.5.22:587')
    s.starttls()
    s.login(msg['From'], password)
    s.sendmail(msg['From'], [msg['To']], msg.as_string())


# URL da página para obter os preços
url = 'https://co.frubana.com/bog/Desechables/contenedores/contenedor-j2-darnel-unid'
headers = {'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/119.0"}

# Obtenha a página da web
response = requests.get(url, headers=headers)
soup = BeautifulSoup(response.content, 'html.parser')

# Data de execução do script
data_execucao = datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')

# Nome do produto
prod_name = item_desc = soup.find('div', class_='productTitle').get_text().strip()

# Encontre todas as tags <div> com a classe "skuItemTiredDetails"
sku_details = soup.find_all("div", class_="skuItemTiredDetails")

# Crie listas vazias para armazenar os dados
min_quantity_list = []
unit_list = []
price_list = []

# Itere pelas tags encontradas
for item in sku_details:
    min_quantity = item.get("data-minquantity")
    unit = item.get("data-unit")

    # Use regex para extrair o preço (removendo caracteres não numéricos)
    price_data = item.get("data-saleprice")
    price = re.sub(r'[^\d.]', '', price_data)  # Remove caracteres não numéricos

    # Converta o preço em um número de ponto flutuante
    price = float(price)

    # Adicione os dados às listas
    min_quantity_list.append(str(min_quantity))  # Trate como texto
    unit_list.append(unit)
    price_list.append(price)

# Crie um DataFrame pandas com os dados
data = {
    'Nome do Produto': prod_name,
    'Min Quantity': min_quantity_list,
    'Unit': unit_list,
    'Price': price_list
}

df = pd.DataFrame(data)
# Adicione a coluna de "Data de Execução"
df['Data de execução'] = data_execucao

# Abra o arquivo Excel existente ou crie um novo se não existir
try:
    book = load_workbook('resultado.xlsx')
except FileNotFoundError:
    book = pd.ExcelWriter('resultado.xlsx', engine='openpyxl').book
    book.create_sheet('Sheet1')

# Se a planilha 'Sheet1' já existir, compare os preços com base na última data de execução para cada quantidade
if 'Sheet1' in book.sheetnames:
    existing_df = pd.read_excel('resultado.xlsx', sheet_name='Sheet1')

    new_rows = []  # Armazena as novas linhas a serem adicionadas
    changed_prices = []  # Armazena as linhas com preços diferentes

    # Compare os preços com base na última data de execução para cada quantidade
    for index, row in df.iterrows():
        min_quantity = row['Min Quantity']
        price = row['Price']

        # Verifique se a quantidade já existe na planilha anterior
        if (existing_df['Min Quantity'] == min_quantity).any():
            last_execution_date = existing_df[existing_df['Min Quantity'] == min_quantity]['Data de execução'].max()
            last_price = existing_df[(existing_df['Min Quantity'] == min_quantity) & (
                        existing_df['Data de execução'] == last_execution_date)]['Price'].values[0]

            if price != last_price:
                # Atualize o preço somente se houver variação
                existing_df = pd.concat([existing_df, row.to_frame().T], ignore_index=True)
                changed_prices.append(row)

        else:
            # Adicione a nova linha se a quantidade e o preço não existirem na planilha
            if not (existing_df['Min Quantity'] == min_quantity).any() and not (existing_df['Price'] == price).any():
                existing_df = pd.concat([existing_df, row.to_frame().T], ignore_index=True)
                new_rows.append(row)

    # Salve as alterações no arquivo Excel
    with pd.ExcelWriter('resultado.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        existing_df.to_excel(writer, 'Sheet1', index=False)

# Após o loop, envie o email apenas se houver novas linhas ou preços diferentes
if new_rows:
    send_new_rows_email(new_rows, prod_name)
    print("Sucesso ao enviar email de novas linhas")

if changed_prices:
    send_changed_prices_email(changed_prices)
    print("Sucesso ao enviar email de preços alterados")

#print(new_rows)
#print(changed_prices)
